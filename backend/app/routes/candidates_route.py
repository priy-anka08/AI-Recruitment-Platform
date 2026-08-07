import os
import io
import uuid
import logging
import requests
import openpyxl
from openpyxl import Workbook
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from app.database.dependencies import get_db
from app.models.candidate import Candidate
from app.models.job import Job
from app.services.ai_resume import (
    generate_interview_questions,
    detect_resume_fraud,
    generate_skill_assessment,
    semantic_rank_candidates,
    generate_offer_recommendation,
    predict_pipeline_success,
)
from app.services.email_service import send_application_status_email

router = APIRouter()
logger = logging.getLogger("uvicorn.error")


@router.get("/")
def get_all_candidates(db: Session = Depends(get_db)):
    return db.query(Candidate).order_by(Candidate.ats_score.desc()).all()


# Manual candidate creation
class ManualCandidateCreate(BaseModel):
    full_name: str
    email: str
    phone: str = ""
    job_id: str
    skills: str = ""
    experience_years: int = 0
    education: str = ""


@router.post("/")
def create_candidate_manually(
    payload: ManualCandidateCreate,
    db: Session = Depends(get_db)
):
    job = db.query(Job).filter(Job.id == payload.job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    existing = db.query(Candidate).filter(Candidate.email == payload.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Candidate with this email already exists")

    candidate = Candidate(
        id=str(uuid.uuid4()),
        full_name=payload.full_name,
        email=payload.email,
        phone=payload.phone,
        job_id=payload.job_id,
        skills=payload.skills,
        experience_years=payload.experience_years,
        education=payload.education,
        ats_score=0.0,
        status="applied"
    )
    db.add(candidate)
    db.commit()
    db.refresh(candidate)
    return candidate


# Export candidates to Excel — MUST be before /{candidate_id} route
@router.get("/export/excel")
def export_candidates_excel(db: Session = Depends(get_db)):
    candidates = db.query(Candidate).order_by(Candidate.ats_score.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    headers = [
        "Full Name", "Email", "Phone", "Job ID", "Skills",
        "Experience (years)", "Education", "ATS Score", "Status", "Resume URL"
    ]
    ws.append(headers)

    for c in candidates:
        ws.append([
            c.full_name,
            c.email,
            c.phone or "",
            c.job_id or "",
            c.skills or "",
            c.experience_years or 0,
            c.education or "",
            c.ats_score or 0,
            c.status or "",
            c.resume_url or "",
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=candidates_export.xlsx"}
    )


# NEW: Semantic Candidate Ranking — MUST be before /{candidate_id} route (uses /rank/ prefix, safe)
@router.get("/rank/{job_id}")
def get_semantic_ranking(job_id: str, db: Session = Depends(get_db)):
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    candidates = db.query(Candidate).filter(Candidate.job_id == job_id).all()
    if not candidates:
        raise HTTPException(status_code=404, detail="No candidates found for this job")

    candidates_data = [
        {
            "candidate_id": c.id,
            "candidate_name": c.full_name,
            "skills": c.skills or "",
            "experience_years": c.experience_years or 0,
            "education": c.education or "",
            "ai_summary": c.ai_summary or "",
        }
        for c in candidates
    ]

    try:
        result = semantic_rank_candidates(candidates_data, job.description or "")
        return result
    except Exception as e:
        logger.error(f"Semantic ranking error: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate semantic ranking")


# NEW: AI Skill Assessment Generator — MCQs based on a job's requirements
@router.get("/assessment/{job_id}")
def get_skill_assessment(job_id: str, db: Session = Depends(get_db)):
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    try:
        assessment = generate_skill_assessment(
            job.title, job.description or "", job.skills_required or ""
        )
        return assessment
    except Exception as e:
        logger.error(f"Skill assessment generation error: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate skill assessment")


# Bulk upload candidates from Excel (columns: Name, Email, Phone)
@router.post("/bulk-upload/{job_id}")
async def bulk_upload_candidates(
    job_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are supported")

    file_bytes = await file.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header row

    added = 0
    skipped = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue

        name = str(row[0]).strip() if len(row) > 0 and row[0] else None
        email = str(row[1]).strip() if len(row) > 1 and row[1] else None
        phone = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        if not name or not email:
            errors.append(f"Row {idx}: Name or Email missing")
            skipped += 1
            continue

        existing = db.query(Candidate).filter(Candidate.email == email).first()
        if existing:
            skipped += 1
            continue

        candidate = Candidate(
            id=str(uuid.uuid4()),
            full_name=name,
            email=email,
            phone=phone,
            job_id=job_id,
            skills="",
            experience_years=0,
            education="",
            ats_score=0.0,
            status="applied"
        )
        db.add(candidate)
        added += 1

    db.commit()

    return {
        "message": "Bulk upload completed",
        "added": added,
        "skipped": skipped,
        "errors": errors
    }


@router.get("/{candidate_id}")
def get_candidate(candidate_id: str, db: Session = Depends(get_db)):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    return candidate


# View / Download resume
@router.get("/{candidate_id}/resume")
def get_resume(candidate_id: str, db: Session = Depends(get_db)):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    if not candidate.resume_url:
        raise HTTPException(status_code=404, detail="Resume not uploaded")
    return {"resume_url": candidate.resume_url}


# AI Interview Question Generator — on-demand, based on resume + job
@router.get("/{candidate_id}/interview-questions")
def get_interview_questions(candidate_id: str, db: Session = Depends(get_db)):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    if not candidate.resume_text:
        raise HTTPException(status_code=400, detail="No resume text available for this candidate")

    job = db.query(Job).filter(Job.id == candidate.job_id).first()
    job_title = job.title if job else ""
    job_description = job.description if job else ""

    try:
        questions = generate_interview_questions(candidate.resume_text, job_title, job_description)
        return questions
    except Exception as e:
        logger.error(f"Interview question generation error: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate interview questions")


# NEW: AI Resume Fraud Detection — on-demand, per candidate
@router.get("/{candidate_id}/fraud-check")
def get_fraud_check(candidate_id: str, db: Session = Depends(get_db)):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    if not candidate.resume_text:
        raise HTTPException(status_code=400, detail="No resume text available for this candidate")

    try:
        result = detect_resume_fraud(candidate.resume_text)
        return result
    except Exception as e:
        logger.error(f"Fraud detection error: {e}")
        raise HTTPException(status_code=500, detail="Failed to run fraud check")


# NEW: AI Offer Recommendation — suggests a salary offer based on candidate profile + job budget
@router.get("/{candidate_id}/offer-recommendation")
def get_offer_recommendation(candidate_id: str, db: Session = Depends(get_db)):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    job = db.query(Job).filter(Job.id == candidate.job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found for this candidate")

    candidate_data = {
        "full_name": candidate.full_name,
        "skills": candidate.skills or "",
        "experience_years": candidate.experience_years or 0,
        "education": candidate.education or "",
        "ats_score": candidate.ats_score or 0,
    }
    job_data = {
        "title": job.title,
        "skills_required": job.skills_required or "",
        "experience_min": job.experience_min or 0,
        "experience_max": job.experience_max or 0,
        "salary_min": job.salary_min or 0,
        "salary_max": job.salary_max or 0,
    }

    try:
        result = generate_offer_recommendation(candidate_data, job_data)
        return result
    except Exception as e:
        logger.error(f"Offer recommendation error: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate offer recommendation")


# NEW: AI Candidate Pipeline Prediction — estimates interview success / offer acceptance likelihood
@router.get("/{candidate_id}/pipeline-prediction")
def get_pipeline_prediction(candidate_id: str, db: Session = Depends(get_db)):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    candidate_data = {
        "full_name": candidate.full_name,
        "status": candidate.status or "",
        "ats_score": candidate.ats_score or 0,
        "skills": candidate.skills or "",
        "matched_skills": candidate.matched_skills or "",
        "missing_skills": candidate.missing_skills or "",
        "experience_years": candidate.experience_years or 0,
        "recommendation_label": candidate.recommendation_label or "",
    }

    try:
        result = predict_pipeline_success(candidate_data)
        return result
    except Exception as e:
        logger.error(f"Pipeline prediction error: {e}")
        raise HTTPException(status_code=500, detail="Failed to generate pipeline prediction")


@router.put("/{candidate_id}/status")
def update_status(
    candidate_id: str,
    status: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    candidate = db.query(Candidate).filter(Candidate.id == candidate_id).first()
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    allowed_statuses = [
        "applied", "under_review", "screened", "shortlisted",
        "interview_scheduled", "technical_round", "hr_round",
        "selected", "rejected", "joined"
    ]

    if status not in allowed_statuses:
        raise HTTPException(status_code=400, detail="Invalid status")

    candidate.status = status
    db.commit()
    db.refresh(candidate)

    background_tasks.add_task(
        send_application_status_email,
        candidate.email,
        candidate.full_name,
        status
    )

    return {"message": f"Status updated to {status}", "candidate": candidate}


@router.get("/status/{status}")
def get_by_status(status: str, db: Session = Depends(get_db)):
    return db.query(Candidate).filter(Candidate.status == status).all()